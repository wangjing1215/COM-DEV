from openpyxl import Workbook
from openpyxl.styles import Color, Font, Alignment, NamedStyle, PatternFill, Border, Side

# 然后如下设置：
# 设置表头字体居中
font = Font(name=u'宋体', bold=True)
align = Alignment(horizontal='center', vertical='center')

# ‘dashDot’,‘dashDotDot’,‘dashed’,‘dotted’,‘double’,‘hair’,‘medium’,‘mediumDashDot’,‘mediumDashDotDot’,‘mediumDashed’,‘slantDashDot’,‘thick’,‘thin’

border = Border(left=Side(border_style="thin", color='272727'),
                right=Side(border_style="thin",  color='272727'),
                top=Side(border_style="thin",  color='272727'),
                bottom=Side(border_style="thin",  color='272727'),
                diagonal=Side(border_style="thin", color='272727'),
                diagonal_direction=0,
                outline=Side(border_style="thin",  color='272727'),
                vertical=Side(border_style="thin", color='272727'),
                horizontal=Side(border_style="thin", color='272727'))

current_col = 1
current_row = 1
xlsxBook = Workbook()
workSheet = xlsxBook.create_sheet('sheet', 0)
workSheet.cell(1, 1).value = '序号'
workSheet.cell(1, 1).alignment = align
workSheet.cell(1, 1).fill = PatternFill('solid', fgColor='F0F8FF')
workSheet.cell(1, 1).border = border
workSheet.merge_cells(start_row=1, end_row=2, start_column=1, end_column=1)


workSheet.cell(1, 2).value = '模块'
workSheet.cell(1, 2).alignment = align
workSheet.cell(1, 2).fill = PatternFill('solid', fgColor='E0FFFF')
workSheet.cell(1, 2).border = border
workSheet.merge_cells(start_row=1, end_row=2, start_column=2, end_column=2)

workSheet.cell(1, 3).value = '查找区块'
workSheet.cell(1, 3).alignment = align
workSheet.cell(1, 3).fill = PatternFill('solid', fgColor='F0F8FF')
workSheet.cell(1, 3).border = border
workSheet.merge_cells(start_row=1, end_row=2, start_column=3, end_column=3)

# workSheet.column_dimensions["A"].width = 18

tmp = {"iram0": ['.data', '.bss', '.text'], "iram1": ['.data', '.bss', '.text'], "iram2": ['.data', '.bss', '.text']}
current_col += 2
for k, v in tmp.items():
    current_col += 1
    workSheet.cell(1, current_col).value = 'iram0'
    workSheet.cell(1, current_col).alignment = align
    workSheet.cell(1, current_col).fill = PatternFill('solid', fgColor='F0F8FF')
    workSheet.cell(1, current_col).border = border
    workSheet.merge_cells(start_row=1, end_row=1, start_column=current_col, end_column=current_col + len(v))
    for po, i in enumerate(v):
        workSheet.cell(2, current_col + po).value = i
    workSheet.cell(2, current_col + len(v)).value = '总计'
    workSheet.cell(2, current_col + len(v)).fill = PatternFill('solid', fgColor='D7FFEE')
    workSheet.cell(2, current_col + len(v)).border = border
    current_col += len(v)

current_col += 1
workSheet.cell(1, current_col).value = '区块总计'
workSheet.cell(1, current_col).alignment = align
workSheet.cell(1, current_col).fill = PatternFill('solid', fgColor='F0F8FF')
workSheet.cell(1, current_col).border = border
workSheet.merge_cells(start_row=1, end_row=2, start_column=current_col, end_column=current_col)

current_col += 1
workSheet.cell(1, current_col).value = '总计'
workSheet.cell(1, current_col).alignment = align
workSheet.cell(1, current_col).fill = PatternFill('solid', fgColor='D7FFEE')
workSheet.cell(1, current_col).border = border
workSheet.merge_cells(start_row=1, end_row=2, start_column=current_col, end_column=current_col)

current_row = 2
res = {"power": [".o", ".32", "ew", "eew", "rrwr"], "ux": ["123", "123", "3123"]}
for po, (k, v) in enumerate(res.items()):
    current_row += 1
    workSheet.cell(current_row, 1).value = "{}".format(po + 1)
    workSheet.cell(current_row, 1).alignment = align
    workSheet.merge_cells(start_row=current_row, end_row=current_row + len(v) - 1, start_column=1, end_column=1)

    workSheet.cell(current_row, 2).value = k
    workSheet.cell(current_row, 2).alignment = align
    workSheet.merge_cells(start_row=current_row, end_row=current_row + len(v) - 1, start_column=2, end_column=2)

    workSheet.merge_cells(start_row=current_row, end_row=current_row + len(v) - 1,
                          start_column=current_col, end_column=current_col)

    for pi, i in enumerate(v):
        workSheet.cell(current_row + pi, 3).value = i
    current_row += len(v) - 1


saveXlsx = '表格.xlsx'
xlsxBook.save(saveXlsx)
